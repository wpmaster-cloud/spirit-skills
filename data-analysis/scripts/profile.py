#!/usr/bin/env python3
"""Profile a tabular data file — schema, row/column counts, null %, distinct counts,
value ranges, and sample values.

Zero-dependency for CSV/TSV/JSON/JSONL and (via openpyxl) XLSX. Parquet needs duckdb
or pandas. Streams the file, so it handles large CSVs without loading everything.

Usage:
  profile.py data.csv
  profile.py data.parquet --json
  profile.py data.tsv --format tsv --samples 5
"""
import argparse
import csv
import json
import os
import sys

DISTINCT_CAP = 50000
SAMPLE_KEEP = 5


def detect_format(path, override):
    if override and override != "auto":
        return override
    ext = os.path.splitext(path)[1].lower()
    return {
        ".csv": "csv", ".tsv": "tsv", ".txt": "csv",
        ".json": "json", ".jsonl": "jsonl", ".ndjson": "jsonl",
        ".parquet": "parquet", ".pq": "parquet",
        ".xlsx": "xlsx", ".xlsm": "xlsx",
    }.get(ext, "csv")


def _sqlstr(s):
    return "'" + s.replace("'", "''") + "'"


def iter_rows(path, fmt):
    if fmt in ("csv", "tsv"):
        delim = "\t" if fmt == "tsv" else None
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            if delim is None:
                sample = f.read(8192)
                f.seek(0)
                try:
                    delim = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
                except Exception:  # noqa: BLE001
                    delim = ","
            for row in csv.DictReader(f, delimiter=delim):
                yield row
    elif fmt == "jsonl":
        with open(path, encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                obj = json.loads(line)
                yield obj if isinstance(obj, dict) else {"value": obj}
    elif fmt == "json":
        with open(path, encoding="utf-8", errors="replace") as f:
            data = json.load(f)
        if isinstance(data, dict):
            for key in ("data", "results", "items", "rows", "records"):
                if isinstance(data.get(key), list):
                    data = data[key]
                    break
            else:
                data = [data]
        for obj in data:
            yield obj if isinstance(obj, dict) else {"value": obj}
    elif fmt == "xlsx":
        yield from _iter_xlsx(path)
    elif fmt == "parquet":
        yield from _iter_parquet(path)
    else:
        sys.exit(f"unsupported format: {fmt}")


def _iter_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("xlsx needs openpyxl:  pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return
    header = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header)]
    for r in rows:
        yield {header[i]: (r[i] if i < len(r) else None) for i in range(len(header))}


def _iter_parquet(path):
    try:
        import duckdb
        df = duckdb.sql(f"SELECT * FROM read_parquet({_sqlstr(path)})").df()
    except Exception:  # noqa: BLE001
        try:
            import pandas as pd
            df = pd.read_parquet(path)
        except Exception:  # noqa: BLE001
            sys.exit("parquet needs duckdb or pandas+pyarrow:  pip install duckdb")
    for rec in df.to_dict("records"):
        yield rec


def classify(v):
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, int):
        return "int"
    if isinstance(v, float):
        return "float"
    s = str(v).strip()
    if s == "":
        return "null"
    low = s.lower()
    if low in ("true", "false"):
        return "bool"
    try:
        int(s)
        return "int"
    except ValueError:
        pass
    try:
        float(s)
        return "float"
    except ValueError:
        pass
    if _looks_date(s):
        return "date"
    return "str"


def _looks_date(s):
    if not (8 <= len(s) <= 32):
        return False
    digits = sum(c.isdigit() for c in s)
    seps = sum(c in "-/:T " for c in s)
    return digits >= 4 and seps >= 2


class Col:
    def __init__(self, name):
        self.name = name
        self.n = 0
        self.nulls = 0
        self.types = {}
        self.distinct = set()
        self.capped = False
        self.num_min = None
        self.num_max = None
        self.num_sum = 0.0
        self.num_n = 0
        self.txt_min = None
        self.txt_max = None
        self.samples = []

    def add(self, v):
        self.n += 1
        t = classify(v)
        if t == "null":
            self.nulls += 1
            return
        self.types[t] = self.types.get(t, 0) + 1
        key = str(v)
        if not self.capped:
            self.distinct.add(key)
            if len(self.distinct) > DISTINCT_CAP:
                self.capped = True
                self.distinct = set()
        if t in ("int", "float"):
            try:
                f = float(v)
                self.num_n += 1
                self.num_sum += f
                self.num_min = f if self.num_min is None else min(self.num_min, f)
                self.num_max = f if self.num_max is None else max(self.num_max, f)
            except (TypeError, ValueError):
                pass
        else:
            if self.txt_min is None or key < self.txt_min:
                self.txt_min = key
            if self.txt_max is None or key > self.txt_max:
                self.txt_max = key
        if len(self.samples) < SAMPLE_KEEP and key not in self.samples:
            self.samples.append(key)

    def dominant_type(self):
        if not self.types:
            return "empty"
        has_int = "int" in self.types
        has_float = "float" in self.types
        non_num = {k: v for k, v in self.types.items() if k not in ("int", "float")}
        if (has_int or has_float) and not non_num:
            return "float" if has_float else "int"
        return max(self.types.items(), key=lambda kv: kv[1])[0]

    def summary(self, total):
        d = {
            "column": self.name,
            "type": self.dominant_type(),
            "non_null": self.n - self.nulls,
            "nulls": self.nulls,
            "null_pct": round(100.0 * self.nulls / total, 1) if total else 0.0,
            "distinct": (">{}".format(DISTINCT_CAP) if self.capped else len(self.distinct)),
            "samples": self.samples,
        }
        if self.num_n:
            d["min"] = _fmtnum(self.num_min)
            d["max"] = _fmtnum(self.num_max)
            d["mean"] = _fmtnum(self.num_sum / self.num_n)
        elif self.txt_min is not None:
            d["min"] = _trunc(self.txt_min)
            d["max"] = _trunc(self.txt_max)
        return d


def _fmtnum(x):
    if x is None:
        return None
    if float(x).is_integer():
        return int(x)
    return round(x, 4)


def _trunc(s, n=24):
    s = str(s)
    return s if len(s) <= n else s[: n - 1] + "…"


def main():
    p = argparse.ArgumentParser(description="Profile a tabular data file.")
    p.add_argument("path")
    p.add_argument("--format", default="auto",
                   help="csv|tsv|json|jsonl|parquet|xlsx (default: by extension)")
    p.add_argument("--samples", type=int, default=5)
    p.add_argument("--json", action="store_true", help="emit JSON instead of a table")
    args = p.parse_args()

    global SAMPLE_KEEP
    SAMPLE_KEEP = args.samples

    if not os.path.isfile(args.path):
        sys.exit(f"no such file: {args.path}")
    fmt = detect_format(args.path, args.format)

    cols = {}
    order = []
    total = 0
    for row in iter_rows(args.path, fmt):
        total += 1
        for k in row:
            if k not in cols:
                cols[k] = Col(k)
                order.append(k)
        for k in order:
            cols[k].add(row.get(k))

    size = os.path.getsize(args.path)
    summaries = [cols[k].summary(total) for k in order]

    if args.json:
        print(json.dumps({
            "file": args.path, "format": fmt, "size_bytes": size,
            "rows": total, "columns": len(order), "schema": summaries,
        }, indent=2, default=str))
        return

    print(f"file:    {args.path}")
    print(f"format:  {fmt}   size: {_human(size)}")
    print(f"rows:    {total}   columns: {len(order)}")
    print()
    headers = ["column", "type", "nulls", "null%", "distinct", "min", "max", "mean", "samples"]
    rows_out = []
    for s in summaries:
        rows_out.append([
            _trunc(s["column"], 28),
            s["type"],
            str(s["nulls"]),
            f'{s["null_pct"]}',
            str(s["distinct"]),
            _trunc(s.get("min", ""), 18),
            _trunc(s.get("max", ""), 18),
            str(s.get("mean", "")),
            _trunc(", ".join(s["samples"]), 30),
        ])
    _print_table(headers, rows_out)


def _human(n):
    n = float(n)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if n < 1024 or unit == "TB":
            return f"{n:.0f} {unit}" if unit == "B" else f"{n:.1f} {unit}"
        n /= 1024.0


def _print_table(headers, rows):
    widths = [len(h) for h in headers]
    for r in rows:
        for i, c in enumerate(r):
            widths[i] = max(widths[i], len(c))
    line = "  ".join(h.ljust(widths[i]) for i, h in enumerate(headers))
    print(line)
    print("  ".join("-" * widths[i] for i in range(len(headers))))
    for r in rows:
        print("  ".join(c.ljust(widths[i]) for i, c in enumerate(r)))


if __name__ == "__main__":
    main()
